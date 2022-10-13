#-*- coding: utf-8 -*-
import os
import re
import time
import json
import xlrd
import random
import string
import win32con
import win32gui
import requests
from datetime import datetime
from selenium import webdriver
from selenium.webdriver import Chrome,ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import Chrome
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import selenium.webdriver.support.expected_conditions as EC
import selenium.webdriver.support.ui as ui


class Get_Cookie():
    def __init__(self):
        self.chrome_path = r'.\OwlBrowser-1.1.6.1'  # 浏览器路径 C:\Program Files\Google\Chrome\Application\chromedriver.exe
        self.executable_path = f'{self.chrome_path}\chromedrivers\99\chromedriver.exe'
        self.userfile = "./AutomationProfile"
        self.chrome_option = ChromeOptions()
        self.ip, self.port = '127.0.0.1',9222
        self.proxy_local_server = 'http://127.0.0.1:7890' # 本地代理ip,用于代理提取外网API
        self.proxy_get_url = 'http://xxxxxxxx.com/api/xxxx' # 提取国外ip接口
        self.excel_file_path = r'.\账号.xlsx'
        self.workbook = None
        

    def chrome_params(self, chrome_options):  # 浏览器参数
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')
        chrome_options.add_argument('--enable-extensions')  # 启用扩展
        return chrome_options
        
    def debug_chrome(self):
        self.chrome_option.add_experimental_option('debuggerAddress', '{}:{}'.format(self.ip, self.port))
        self.chrome_option = self.chrome_params(self.chrome_option)
        return self.chrome_option
    
    def start_brower(self, browserVersion, proxy_server=None, glat=None, timezone=None): #启动浏览器
        import uuid
        # 获取mac地址
        address = hex(uuid.getnode())[2:]
        macAddress = ':'.join(address[i:i + 2] for i in range(0, len(address), 2))
        # 获取计算机名
        letters = [i for i in string.ascii_letters]
        digits = [i for i in string.digits]

        computerName = ''.join(
            random.sample(letters, random.randint(5, 10)) + ['-'] + random.sample(digits, random.randint(5, 10)))
        rand_hash = random.randint(1, 99999999)
        lis = [(800, 600), (1024, 768), (1152, 864), (1280, 600), (1280, 720), (1280, 768), (1280, 800), (1280, 960),
               (1280, 1024), (1360, 768), (1366, 768),
               (1400, 1050), (1440, 900), (1600, 900), (1680, 1050), (1920, 1080)]
        screenWidth, screenHeight = random.choice(lis)
        
        platform = random.choice(['Win64'])
        Concurrency = random.randint(1, 8)
        deviceMemory = random.randint(1, 16)
        colorDepth = random.choice([16, 24, 32, 48])
        gtx = [f'GTX {i}' for i in [750, 1050, 1660, 970, 1650, 2060, 2080, 2070]]
        rtx = [f'RTX {i}' for i in [3050, 3060, 3070, 3080, 3090]]
        gpu = random.choice(gtx + rtx)
        gpuRenderer = f"NVIDIA GeForce {gpu} Ti Direct3D11 vs_5_0 ps_5_0&"
        userName = ''.join(random.sample(letters, random.randint(5, 10)) + random.sample(digits, random.randint(5, 10)))
        # glat = 33984+random.randint(-100,100)
        glmcmts = 16384 + random.randint(-100, 100)
        glmvuv = 4095 + random.randint(-50, 50)
        glsw = 2147483647 + random.randint(-100, 100)
        glsbvm = 2147483647 + random.randint(-100, 100)
        gl2mcfuc = 200704 + random.randint(-100, 100)
        gl2mcvuc = 212988 + random.randint(-100, 100)
        gl2mei = 2147483647 + random.randint(-100, 100)
        
        plugins = ' --plugins="plugName=p1&plugDesc=p2&plugPath=p3&mimeType=p4&mimeExt=p5&mimeDesc=p2&,plugName=n1&plugDesc=n2&plugPath=n3&mimeType=n4&mimeExt=n5&mimeDesc=n2&"'
        browser_param = ' --browser-common-param="languages=en-US&disableWebRtc=1&platform={}&computerName={}&macAddress={}&screenWidth={}&screenHeight={}&hardwareConcurrency={}&deviceMemory={}&colorDepth={}&gpuRenderer={}&gpuVendor=Google Inc.&productSub=20030107&vendor=Google Inc.&vendorSub=&protectPorts=all&userName={}&filterGlblob=1&filterGldata=1&filterPixel=1&glmcmts={}&glmvuv={}&glsw={}&glsbvm={}&gl2mcfuc={}&gl2mcvuc={}&gl2mei={}&browserVersion={}"'
        browser_param = browser_param.format(platform, computerName, macAddress, screenWidth, screenHeight, Concurrency,
        deviceMemory,colorDepth, gpuRenderer, userName, glmcmts, glmvuv, glsw, glsbvm, gl2mcfuc,gl2mcvuc, gl2mei, browserVersion)
        start_params = r'cd {} && chrome.exe --user-data-dir="{}" --remote-debugging-port={} --browser-rand-hash={} --disabled-target-blank' + plugins
        if proxy_server: # 代理
            start_params += ' --proxy-server={}'.format(proxy_server)
        if timezone:
            browser_param += f'&timezone={timezone}'
        if glat:
            browser_param += f'&glat={glat}'
        start_params += browser_param
        os.popen(start_params.format(self.chrome_path, self.userfile, self.port, rand_hash))

    def start_brower1(self,ip=None):
        if ip:
            proxy_server = ip
        else:
            proxy_server = None
       
        ua_versions = ['89.0.4109.36', '81.0.4044.138', '81.0.4044.20', '81.0.4044.69', '83.0.4103.14', '83.0.4103.39',
         '84.0.4147.30', '85.0.4183.38', '85.0.4183.83', '85.0.4183.87', '86.0.4240.22', '87.0.4280.20','87.0.4280.87',
        '87.0.4280.88', '88.0.4324.27', '88.0.4324.96', '89.0.4389.23', '90.0.4430.24','91.0.4472.101', '91.0.4472.19',
        '92.0.4515.107', '92.0.4515.43', '93.0.4577.15', '93.0.4577.63','94.0.4606.113', '94.0.4606.41', '94.0.4606.61',
        '95.0.4638.10', '95.0.4638.17', '95.0.4638.54','95.0.4638.69', '96.0.4664.18', '96.0.4664.35', '96.0.4664.45',
        '97.0.4692.20', '97.0.4692.36','97.0.4692.71', '98.0.4758.102', '98.0.4758.48', '98.0.4758.80', '99.0.4844.17',
        '99.0.4844.35','99.0.4844.51']

        ua_version = random.choice(ua_versions) #随机选择版本
        version_num = ua_version.split('.')[0]
        self.start_brower(ua_version, proxy_server)
        capa = DesiredCapabilities.CHROME
        capa["pageLoadStrategy"] = "none"
        self.driver = Chrome(executable_path=self.executable_path.replace('99', version_num),options=self.debug_chrome(),desired_capabilities=capa)
        #self.driver.maximize_window() #最大化窗口
        # titlename = pyautogui.getActiveWindowTitle()
        # hwnd = win32gui.FindWindow(0, titlename)  # 获取句柄
        # left, top, right, bottom = win32gui.GetWindowRect(hwnd)
        # win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 1000, 878, win32con.SWP_SHOWWINDOW)
        
    def stopLoading(self): #停止加载页面
        import win32api
        win32api.keybd_event(27, 0, 0, 0)
        win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

    # 检查元素是否存在
    def check_element_exists(driver, element, condition):
        try:
            if condition == 'class':
                driver.find_element_by_class_name(element)
            elif condition == 'id':
                driver.find_element_by_id(element)
            elif condition == 'xpath':
                driver.find_element_by_xpath(element)
            return True
        except Exception as e:
            return False

    def check_load(self):  # 判断页面是否加载成功
        print('检测页面是否加载成功')
        flug = False #
        for i in range(4):
            # res=self.driver.execute_script("return document.readyState") #.equals("complete")
            # res = self.driver.find_element_by_name('username')
            # res = self.check_element_exists(self.driver, 'username', 'xpath')
            try:
                self.driver.find_element_by_name('username')
                flug = True
                break
            except:
                if i%2==0:
                    self.driver.refresh()
            time.sleep(2)
        print('check_load:',flug)
        return flug
        
    def request_url(self,url): #发送请求
        print('当前请求的url:', url)
        request_count = 0
        flug = False
        check_load_status=False
        while True:
            if request_count > 2: break
            try:
                self.driver.get(url=url)
                time.sleep(2)
                title = ''.join(re.findall('<title.*?>(.*?)</title>',self.driver.page_source))
                print('title:',title)
                if '404' in title or 'Not Found'.lower() in title.lower() or 'show_recurrent_error_paragraph' in self.driver.page_source or '隐私设置错误' in title:
                    self.response_data.append((2, '网站失效', datetime.now().strftime('%Y-%m-%d'), data['id']))
                    break
                if self.check_load(): #
                    check_load_status=True
                    self.stopLoading()
                    flug = True
                    break
                request_count += 1
            except Exception as ex:
                request_count += 1
                print('ex:', ex)
        
        return flug
        
    def clear_cache(self):  # 清除浏览器缓存
        print('清除浏览器缓存')
        time.sleep(5)
        js0 = """()=>{document.querySelector("body > settings-ui").shadowRoot.querySelector("#main").shadowRoot.querySelector("settings-basic-page").shadowRoot.querySelector("#basicPage > settings-section:nth-child(7) > settings-privacy-page").shadowRoot.querySelector("settings-clear-browsing-data-dialog").shadowRoot.querySelector("#clearBrowsingDataConfirm").click()}"""
        try:
            self.driver.get('chrome://settings/clearBrowserData')
            self.driver.execute_script(js0)
            clearButton = self.driver.execute_script(
                'return document.querySelector("body > settings-ui").shadowRoot.querySelector("#main").shadowRoot.querySelector("settings-basic-page").shadowRoot.querySelector("#basicPage > settings-section:nth-child(7) > settings-privacy-page").shadowRoot.querySelector("settings-clear-browsing-data-dialog").shadowRoot.querySelector("#clearBrowsingDataConfirm")')
            clearButton.click()
        except:pass
        
    def redExcelData(self): #读取Excel数据
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_file_path)
        sheet = wb.active
        rows_data = list(sheet.rows)
        titles = [title.value for title in rows_data.pop(0)]
        all_row_dict = []
        for a_row in rows_data:
            the_row_data = [cell.value for cell in a_row]
            row_dict = dict(zip(titles, the_row_data))
            all_row_dict.append(row_dict)
        self.workbook = wb
        self.sheet = sheet
        return all_row_dict

    def getCookies(self): # 获取cookies
        cookie_list = self.driver.get_cookies()
        cookies = ";".join([item["name"] + "=" + item["value"] + "" for item in cookie_list])
        return cookies

    def get_proxy(self): #获取代理
        proxies = {}
        proxies[self.proxy_local_server.split('://')[0]] = self.proxy_local_server.split('://')[1]
        response = requests.get(self.proxy_get_url, proxies=proxies)
        if(response.status_code == 200):
            data = json.loads(response.text)
            if(data.get('code') == 0):
                print('获取代理ip成功',data.get('data'))
                return "socks5://"+data.get('data')[0]
            else:
                print('获取代理ip失败',data.get('msg'))
                return None
        else:
            print("提取境外代理ip失败，请检查本地代理ip是否正常")
            return None
    
    # 输入用户名密码
    def loginTiktok(self,username,password):
        try:
            self.driver.find_element_by_name('username').click()
            self.driver.find_element_by_name('username').clear()
            time.sleep(1)
            self.driver.find_element_by_name('username').send_keys(username)
            time.sleep(2)
            self.driver.find_element_by_css_selector('input[autocomplete="new-password"]').click()
            self.driver.find_element_by_css_selector('input[autocomplete="new-password"]').clear()
            self.driver.find_element_by_css_selector('input[autocomplete="new-password"]').send_keys(password)
            print('账号登入')
            self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/form/button').click()
            # /html/body/div[2]/div/div[2]/div[1]/form/div[5] 出错提示
            # 判断是否进入主页面，登录成功，如果没有进入主页面，则继续登录，超过一定次数，则跳出循环
            res = wait(self.driver, 15, 0.5).until(
            # 搜索框按钮
                EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[1]/div/div[1]/div/form/button"))
            )
            return True
        except Exception as e:
            print("登录错误")
            return False

    def main(self):
        data = self.redExcelData()
        for (index,item) in enumerate(data):
            if((item.get('cookie') is not None) or (item.get('账号') is None) or (item.get('密码') is None)):
                print("第{}行cookie已存在/账号密码为空，跳过".format(index+2))
                continue

            ip = self.get_proxy()
            # ip = 'socks5://182.160.15.888:6666'
            self.start_brower1(ip)
        
            url = 'https://www.tiktok.com/login/phone-or-email/email'
            self.response_data = []
            try:
                result = self.request_url(url)
                print('result:', result)
                for i in range(3):
                    login_result = self.loginTiktok(item.get('账号'),item.get('密码'))
                    # self.driver.implicitly_wait(10)
                    if(login_result):
                        print('{}登录成功'.format(item.get('账号')))
                        self.sheet.cell(index+2,4).value = self.getCookies()
                        break
                    else:
                        self.driver.refresh()
                        time.sleep(2)
                if(login_result == False):
                    print("{}账号登录失败".format(item.get('账号')))
                    self.sheet.cell(index+2,3).value = '登录失败'   
                self.workbook.save(self.excel_file_path)
                self.clear_cache()
                self.skill_brower()
            except Exception as e:
                print(e)
                self.skill_brower()
            
    def skill_brower(self): #关掉浏览器进程
        print('关掉浏览器进程')
        try:
            self.driver.close()
            self.driver.quit()
        except:pass
        os.system("taskkill /f /im chrome.exe")
        os.system('taskkill /f /im chromedriver.exe')

if __name__ == '__main__':
    gtk = Get_Cookie()
    gtk.main()
   
    
    
    










